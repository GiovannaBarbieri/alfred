import csv
import io
from datetime import date, datetime
from typing import Literal

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.api.routes.filter_utils import build_filter_clause
from app.api.routes.reports import (
    GROUP_CONFIG,
    ReportGroup,
    get_project_comparison,
    get_project_evolution,
)
from app.db import get_connection

router = APIRouter()

ExportKind = Literal["consolidated", "report"]


@router.get("/consolidated.csv")
def export_consolidated_csv(
    start_date: date | None = Query(None, alias="startDate"),
    end_date: date | None = Query(None, alias="endDate"),
    user: str | None = None,
    epic_id: str | None = Query(None, alias="epicId"),
    category: str | None = None,
    import_id: int | None = Query(None, alias="importId"),
) -> StreamingResponse:
    where_sql, params = build_filter_clause(
        start_date=start_date,
        end_date=end_date,
        user=user,
        epic_id=epic_id,
        category=category,
        import_id=import_id,
    )

    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                    i.id AS importacao_id,
                    i.nome_arquivo,
                    l.id_lancamento,
                    l.data_hora_cadastro,
                    l.login_usuario,
                    l.duracao_original,
                    l.duracao_segundos,
                    ROUND((l.duracao_segundos::numeric / 3600), 2) AS duracao_horas,
                    l.id_epic,
                    l.titulo_epic,
                    l.id_feat,
                    l.titulo_feat,
                    l.id_pbi,
                    l.titulo_pbi,
                    l.id_task,
                    l.titulo_task,
                    COALESCE(c.nome, 'Nao classificado') AS categoria,
                    COALESCE(s.nome, 'Nao classificado') AS subcategoria,
                    l.status_validacao,
                    l.status_classificacao
                FROM lancamentos_horas l
                JOIN importacoes i ON i.id = l.importacao_id
                LEFT JOIN categorias c ON c.id = l.categoria_id
                LEFT JOIN subcategorias s ON s.id = l.subcategoria_id
                {where_sql}
                ORDER BY l.data_hora_cadastro, l.id
                """,
                params,
            )
            rows = cursor.fetchall()

    headers = [
        "ImportacaoId",
        "Arquivo",
        "IdLancamento",
        "DataHoraCadastro",
        "LoginUsuario",
        "Duracao",
        "DuracaoCalculada",
        "DuracaoHorasDecimal",
        "IdEpic",
        "TituloEpic",
        "IdFeat",
        "TituloFeat",
        "IdPBI",
        "TituloPBI",
        "IdTask",
        "TituloTask",
        "Categoria",
        "Subcategoria",
        "StatusValidacao",
        "StatusClassificacao",
    ]
    csv_rows = [
        [
            row["importacao_id"],
            row["nome_arquivo"],
            row["id_lancamento"],
            row["data_hora_cadastro"].isoformat(sep=" "),
            row["login_usuario"],
            row["duracao_original"],
            seconds_to_hhmmss(int(row["duracao_segundos"])),
            row["duracao_horas"],
            row["id_epic"],
            row["titulo_epic"],
            row["id_feat"],
            row["titulo_feat"],
            row["id_pbi"],
            row["titulo_pbi"],
            row["id_task"],
            row["titulo_task"],
            row["categoria"],
            row["subcategoria"],
            row["status_validacao"],
            row["status_classificacao"],
        ]
        for row in rows
    ]

    return _csv_response("base-consolidada.csv", headers, csv_rows)


@router.get("/report.csv")
def export_report_csv(
    group_by: ReportGroup = Query(..., alias="groupBy"),
    start_date: date | None = Query(None, alias="startDate"),
    end_date: date | None = Query(None, alias="endDate"),
    user: str | None = None,
    epic_id: str | None = Query(None, alias="epicId"),
    category: str | None = None,
    import_id: int | None = Query(None, alias="importId"),
) -> StreamingResponse:
    config = GROUP_CONFIG.get(group_by)
    if not config:
        raise HTTPException(status_code=400, detail="Agrupamento invalido.")

    where_sql, params = build_filter_clause(
        start_date=start_date,
        end_date=end_date,
        user=user,
        epic_id=epic_id,
        category=category,
        import_id=import_id,
    )
    key_expr = config["key"]
    label_expr = config["label"]

    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                WITH grouped AS (
                    SELECT
                        {key_expr} AS chave,
                        {label_expr} AS descricao,
                        SUM(l.duracao_segundos) AS total_segundos,
                        COUNT(*) AS total_lancamentos
                    FROM lancamentos_horas l
                    LEFT JOIN categorias c ON c.id = l.categoria_id
                    LEFT JOIN subcategorias s ON s.id = l.subcategoria_id
                    {where_sql}
                    GROUP BY {key_expr}, {label_expr}
                ),
                totals AS (
                    SELECT COALESCE(SUM(total_segundos), 0) AS total_geral
                    FROM grouped
                )
                SELECT
                    grouped.chave,
                    grouped.descricao,
                    grouped.total_segundos,
                    ROUND((grouped.total_segundos::numeric / 3600), 2) AS total_horas,
                    grouped.total_lancamentos,
                    CASE
                        WHEN totals.total_geral = 0 THEN 0
                        ELSE ROUND((grouped.total_segundos::numeric / totals.total_geral) * 100, 2)
                    END AS percentual
                FROM grouped
                CROSS JOIN totals
                ORDER BY grouped.total_segundos DESC, grouped.descricao
                """,
                params,
            )
            rows = cursor.fetchall()

    headers = ["Chave", "Descricao", "TotalSegundos", "TotalDuracao", "TotalHorasDecimal", "TotalLancamentos", "Percentual"]
    csv_rows = [
        [
            row["chave"],
            row["descricao"],
            row["total_segundos"],
            seconds_to_hhmmss(int(row["total_segundos"])),
            row["total_horas"],
            row["total_lancamentos"],
            row["percentual"],
        ]
        for row in rows
    ]
    return _csv_response(f"relatorio-{group_by}.csv", headers, csv_rows)


@router.get("/project-analysis.xlsx")
def export_project_analysis_xlsx(
    import_id: int = Query(..., alias="importId"),
    user: str | None = Query(None),
) -> StreamingResponse:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute("SELECT nome_arquivo FROM importacoes WHERE id = %s", [import_id])
            import_row = cursor.fetchone()
            if not import_row:
                raise HTTPException(status_code=404, detail="Importacao nao encontrada.")

            sheets = [
                (
                    "Diario_Total",
                    """
                    SELECT
                        l.data_hora_cadastro::date AS periodo,
                        NULL AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    WHERE l.importacao_id = %s
                    GROUP BY l.data_hora_cadastro::date
                    ORDER BY periodo
                    """,
                ),
                (
                    "Dia_Colaborador",
                    """
                    SELECT
                        l.data_hora_cadastro::date AS periodo,
                        l.login_usuario AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    WHERE l.importacao_id = %s
                    GROUP BY l.data_hora_cadastro::date, l.login_usuario
                    ORDER BY periodo, serie
                    """,
                ),
                (
                    "Semana_Colaborador",
                    """
                    SELECT
                        DATE_TRUNC('week', l.data_hora_cadastro)::date AS periodo,
                        l.login_usuario AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    WHERE l.importacao_id = %s
                    GROUP BY DATE_TRUNC('week', l.data_hora_cadastro)::date, l.login_usuario
                    ORDER BY periodo, serie
                    """,
                ),
                (
                    "Dia_Categoria",
                    """
                    SELECT
                        l.data_hora_cadastro::date AS periodo,
                        COALESCE(c.nome, 'Nao classificado') AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    LEFT JOIN categorias c ON c.id = l.categoria_id
                    WHERE l.importacao_id = %s
                    GROUP BY l.data_hora_cadastro::date, COALESCE(c.nome, 'Nao classificado')
                    ORDER BY periodo, serie
                    """,
                ),
                (
                    "Semana_Categoria",
                    """
                    SELECT
                        DATE_TRUNC('week', l.data_hora_cadastro)::date AS periodo,
                        COALESCE(c.nome, 'Nao classificado') AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    LEFT JOIN categorias c ON c.id = l.categoria_id
                    WHERE l.importacao_id = %s
                    GROUP BY DATE_TRUNC('week', l.data_hora_cadastro)::date, COALESCE(c.nome, 'Nao classificado')
                    ORDER BY periodo, serie
                    """,
                ),
                (
                    "Mes_Categoria",
                    """
                    SELECT
                        DATE_TRUNC('month', l.data_hora_cadastro)::date AS periodo,
                        COALESCE(c.nome, 'Nao classificado') AS serie,
                        SUM(l.duracao_segundos) AS total_segundos,
                        ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas
                    FROM lancamentos_horas l
                    LEFT JOIN categorias c ON c.id = l.categoria_id
                    WHERE l.importacao_id = %s
                    GROUP BY DATE_TRUNC('month', l.data_hora_cadastro)::date, COALESCE(c.nome, 'Nao classificado')
                    ORDER BY periodo, serie
                    """,
                ),
            ]

            workbook = Workbook()
            index_sheet = workbook.active
            index_sheet.title = "Indice"

            for title, sql in sheets:
                cursor.execute(sql, [import_id])
                rows = cursor.fetchall()
                max_series = None if title == "Diario_Total" else 5
                _append_timeline_chart_sheet(
                    workbook.create_sheet(title),
                    title.replace("_", " "),
                    rows,
                    max_series=max_series,
                )

            task_where = "WHERE l.importacao_id = %s"
            task_params: list = [import_id]
            if user:
                task_where += " AND l.login_usuario = %s"
                task_params.append(user)

            cursor.execute(
                f"""
                SELECT
                    l.id_task,
                    l.titulo_task,
                    l.login_usuario,
                    COALESCE(c.nome, 'Nao classificado') AS categoria,
                    COALESCE(s.nome, 'Nao classificado') AS subcategoria,
                    SUM(l.duracao_segundos) AS total_segundos,
                    ROUND((SUM(l.duracao_segundos)::numeric / 3600), 2) AS total_horas,
                    COUNT(*) AS total_lancamentos
                FROM lancamentos_horas l
                LEFT JOIN categorias c ON c.id = l.categoria_id
                LEFT JOIN subcategorias s ON s.id = l.subcategoria_id
                {task_where}
                GROUP BY l.id_task, l.titulo_task, l.login_usuario, c.nome, s.nome
                ORDER BY total_segundos DESC, l.titulo_task
                """,
                task_params,
            )
            task_rows = cursor.fetchall()
            _append_sheet(
                workbook.create_sheet("Tasks"),
                [
                    "IdTask",
                    "TituloTask",
                    "LoginUsuario",
                    "Categoria",
                    "Subcategoria",
                    "TotalSegundos",
                    "TotalDuracao",
                    "TotalHorasDecimal",
                    "TotalLancamentos",
                ],
                [
                    [
                        row["id_task"],
                        row["titulo_task"],
                        row["login_usuario"],
                        row["categoria"],
                        row["subcategoria"],
                        int(row["total_segundos"]),
                        seconds_to_hhmmss(int(row["total_segundos"])),
                        float(row["total_horas"]),
                        row["total_lancamentos"],
                    ]
                    for row in task_rows
                ],
            )
            _append_index_sheet(index_sheet, workbook.worksheets, import_row["nome_arquivo"], user)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"analise-projeto-{import_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/project-comparison.xlsx")
def export_project_comparison_xlsx(import_ids: list[int] = Query(..., alias="importIds")) -> StreamingResponse:
    comparison = get_project_comparison(import_ids)
    projects = comparison["projects"]
    if len(projects) < 2:
        raise HTTPException(status_code=400, detail="Selecione pelo menos dois projetos para exportar.")

    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Indice"

    _append_sheet(
        workbook.create_sheet("Resumo"),
        ["Campo", "Valor"],
        [
            ["ProjetosComparados", comparison["summary"]["projectsCount"]],
            ["TotalHoras", comparison["summary"]["totalHours"]],
            ["TotalRegistros", comparison["summary"]["recordsCount"]],
            ["PendenciasAbertas", comparison["summary"]["openPendings"]],
            ["GeradoEm", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ],
    )
    _append_sheet(
        workbook.create_sheet("Comparativo"),
        [
            "ImportacaoId",
            "Projeto",
            "Arquivo",
            "DataImportacao",
            "Status",
            "TotalSegundos",
            "TotalHoras",
            "Registros",
            "Colaboradores",
            "Tasks",
            "PendenciasAbertas",
            "TaxaPendencias",
            "NivelAtencao",
            "PendenciasRevisadas",
            "PendenciasIgnoradas",
            "CategoriaPrincipal",
            "CategoriaPrincipalPercentual",
            "ColaboradorPrincipal",
            "ColaboradorPrincipalPercentual",
        ],
        [
            [
                project["importId"],
                project["projectName"],
                project["filename"],
                project["importedAt"],
                project["status"],
                project["totalSeconds"],
                project["totalHours"],
                project["recordsCount"],
                project["collaboratorsCount"],
                project["tasksCount"],
                project["openPendings"],
                project["pendingRate"],
                project["attentionLabel"],
                project["reviewedPendings"],
                project["ignoredPendings"],
                project["topCategory"],
                project["topCategoryPercentage"],
                project["topCollaborator"],
                project["topCollaboratorPercentage"],
            ]
            for project in projects
        ],
    )

    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Horas"),
        "Comparativo de horas",
        "Horas",
        [[project["projectName"], project["totalHours"]] for project in projects],
    )
    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Registros"),
        "Comparativo de registros",
        "Registros",
        [[project["projectName"], project["recordsCount"]] for project in projects],
    )
    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Pendencias"),
        "Comparativo de pendencias abertas",
        "Pendencias abertas",
        [[project["projectName"], project["openPendings"]] for project in projects],
    )
    _append_comparison_index_sheet(index_sheet, workbook.worksheets)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="comparativo-projetos.xlsx"'},
    )


@router.get("/project-evolution.xlsx")
def export_project_evolution_xlsx(project_name: str = Query(..., alias="projectName")) -> StreamingResponse:
    evolution = get_project_evolution(project_name)
    points = evolution["points"]

    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Indice"

    _append_sheet(
        workbook.create_sheet("Resumo"),
        ["Campo", "Valor"],
        [
            ["Projeto", evolution["projectName"]],
            ["Importacoes", evolution["importsCount"]],
            ["PrimeiraImportacao", evolution["firstImportedAt"]],
            ["UltimaImportacao", evolution["latestImportedAt"]],
            ["VariacaoHoras", evolution["summary"]["hoursDelta"]],
            ["VariacaoRegistros", evolution["summary"]["recordsDelta"]],
            ["VariacaoPendencias", evolution["summary"]["pendingsDelta"]],
            ["AtencaoInicial", evolution["summary"]["firstAttention"]],
            ["AtencaoAtual", evolution["summary"]["latestAttention"]],
            ["Tendencia", evolution["summary"]["trendLabel"]],
            ["GeradoEm", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ],
    )
    _append_sheet(
        workbook.create_sheet("Historico"),
        [
            "ImportacaoId",
            "Arquivo",
            "DataImportacao",
            "Status",
            "TotalHoras",
            "VariacaoHoras",
            "Registros",
            "VariacaoRegistros",
            "PendenciasAbertas",
            "VariacaoPendencias",
            "TaxaPendencias",
            "NivelAtencao",
            "AtencaoAlterada",
        ],
        [
            [
                point["importId"],
                point["filename"],
                point["importedAt"],
                point["status"],
                point["totalHours"],
                point["hoursDelta"],
                point["recordsCount"],
                point["recordsDelta"],
                point["openPendings"],
                point["pendingsDelta"],
                point["pendingRate"],
                point["attentionLabel"],
                "Sim" if point["attentionChanged"] else "Nao",
            ]
            for point in points
        ],
    )
    _append_sheet(
        workbook.create_sheet("Alertas"),
        ["Prioridade", "Titulo", "Motivo", "Acao", "Origem"],
        [
            [
                insight["priority"],
                insight["title"],
                insight["reason"],
                insight["action"],
                insight["source"],
            ]
            for insight in evolution["insights"]
        ],
    )
    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Horas"),
        "Evolucao de horas",
        "Horas",
        [[_evolution_point_label(index, point), point["totalHours"]] for index, point in enumerate(points)],
    )
    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Pendencias"),
        "Evolucao de pendencias",
        "Pendencias abertas",
        [[_evolution_point_label(index, point), point["openPendings"]] for index, point in enumerate(points)],
    )
    _append_comparison_chart_sheet(
        workbook.create_sheet("Grafico_Registros"),
        "Evolucao de registros",
        "Registros",
        [[_evolution_point_label(index, point), point["recordsCount"]] for index, point in enumerate(points)],
    )
    _append_evolution_index_sheet(index_sheet, workbook.worksheets, evolution["projectName"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="evolucao-projeto.xlsx"'},
    )


def _csv_response(filename: str, headers: list[str], rows: list[list]) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _append_sheet(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="E8F1FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    _format_sheet_numbers(sheet)
    _autosize_columns(sheet)


def _append_index_sheet(index_sheet, worksheets, filename: str, user: str | None) -> None:
    index_sheet.sheet_view.showGridLines = False
    index_sheet.append(["Analise operacional de horas"])
    index_sheet.append(["Arquivo", filename])
    index_sheet.append(["Filtro colaborador", user or "Todos"])
    index_sheet.append(["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    index_sheet.append([])
    index_sheet.append(["Aba", "Conteudo"])

    header_fill = PatternFill("solid", fgColor="123C69")
    title_cell = index_sheet["A1"]
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = header_fill
    index_sheet.merge_cells("A1:B1")

    for cell in index_sheet[6]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F1FF")

    descriptions = {
        "Diario_Total": "Linha do tempo diaria do total do projeto.",
        "Dia_Colaborador": "Horas por dia e colaborador.",
        "Semana_Colaborador": "Horas por semana e colaborador.",
        "Dia_Categoria": "Horas por dia e categoria.",
        "Semana_Categoria": "Horas por semana e categoria.",
        "Mes_Categoria": "Horas mensais por categoria.",
        "Tasks": "Tasks agrupadas com duracao e classificacao.",
    }
    for sheet in worksheets:
        if sheet.title == "Indice":
            continue
        row_number = index_sheet.max_row + 1
        index_sheet.append([sheet.title, descriptions.get(sheet.title, "Dados da analise.")])
        link_cell = index_sheet.cell(row=row_number, column=1)
        link_cell.hyperlink = f"#'{sheet.title}'!A1"
        link_cell.style = "Hyperlink"

    index_sheet.freeze_panes = "A7"
    index_sheet.auto_filter.ref = f"A6:B{index_sheet.max_row}"
    _autosize_columns(index_sheet)


def _append_comparison_index_sheet(index_sheet, worksheets) -> None:
    index_sheet.sheet_view.showGridLines = False
    index_sheet.append(["Comparativo operacional de projetos"])
    index_sheet.append(["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    index_sheet.append([])
    index_sheet.append(["Aba", "Conteudo"])

    header_fill = PatternFill("solid", fgColor="123C69")
    title_cell = index_sheet["A1"]
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = header_fill
    index_sheet.merge_cells("A1:B1")

    for cell in index_sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F1FF")

    descriptions = {
        "Resumo": "Indicadores consolidados dos projetos selecionados.",
        "Comparativo": "Tabela completa com metricas por projeto.",
        "Grafico_Horas": "Grafico de barras comparando horas totais.",
        "Grafico_Registros": "Grafico de barras comparando quantidade de registros.",
        "Grafico_Pendencias": "Grafico de barras comparando pendencias abertas.",
    }
    for sheet in worksheets:
        if sheet.title == "Indice":
            continue
        row_number = index_sheet.max_row + 1
        index_sheet.append([sheet.title, descriptions.get(sheet.title, "Dados do comparativo.")])
        link_cell = index_sheet.cell(row=row_number, column=1)
        link_cell.hyperlink = f"#'{sheet.title}'!A1"
        link_cell.style = "Hyperlink"

    index_sheet.freeze_panes = "A5"
    index_sheet.auto_filter.ref = f"A4:B{index_sheet.max_row}"
    _autosize_columns(index_sheet)


def _append_evolution_index_sheet(index_sheet, worksheets, project_name: str) -> None:
    index_sheet.sheet_view.showGridLines = False
    index_sheet.append(["Evolucao operacional do projeto"])
    index_sheet.append(["Projeto", project_name])
    index_sheet.append(["Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    index_sheet.append([])
    index_sheet.append(["Aba", "Conteudo"])

    header_fill = PatternFill("solid", fgColor="123C69")
    title_cell = index_sheet["A1"]
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = header_fill
    index_sheet.merge_cells("A1:B1")

    for cell in index_sheet[5]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F1FF")

    descriptions = {
        "Resumo": "Indicadores de variacao entre a primeira e a ultima importacao.",
        "Historico": "Historico completo das importacoes do projeto.",
        "Alertas": "Alertas operacionais gerados pela evolucao.",
        "Grafico_Horas": "Grafico de barras da evolucao das horas.",
        "Grafico_Pendencias": "Grafico de barras da evolucao das pendencias abertas.",
        "Grafico_Registros": "Grafico de barras da evolucao de registros.",
    }
    for sheet in worksheets:
        if sheet.title == "Indice":
            continue
        row_number = index_sheet.max_row + 1
        index_sheet.append([sheet.title, descriptions.get(sheet.title, "Dados da evolucao.")])
        link_cell = index_sheet.cell(row=row_number, column=1)
        link_cell.hyperlink = f"#'{sheet.title}'!A1"
        link_cell.style = "Hyperlink"

    index_sheet.freeze_panes = "A6"
    index_sheet.auto_filter.ref = f"A5:B{index_sheet.max_row}"
    _autosize_columns(index_sheet)


def _format_sheet_numbers(sheet) -> None:
    decimal_headers = {
        "TotalHorasDecimal",
        "TotalHoras",
        "Percentual",
    }
    for column_index, header_cell in enumerate(sheet[1], start=1):
        header = str(header_cell.value or "")
        if header in decimal_headers or column_index > 1 and sheet.title != "Resumo":
            for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                cell = row[0]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 52)


def _append_timeline_chart_sheet(sheet, title: str, rows: list[dict], max_series: int | None = None) -> None:
    periods = sorted({row["periodo"].isoformat() for row in rows})
    series_names = sorted(
        {row["serie"] or "Total" for row in rows},
        key=lambda series: (
            -sum(float(row["total_horas"]) for row in rows if (row["serie"] or "Total") == series),
            series,
        ),
    )
    if max_series is not None:
        series_names = series_names[:max_series]
    values = {
        (row["periodo"].isoformat(), row["serie"] or "Total"): float(row["total_horas"])
        for row in rows
    }

    _append_sheet(
        sheet,
        ["Periodo", *series_names],
        [[date.fromisoformat(period).strftime("%d/%m/%Y"), *[values.get((period, series), 0) for series in series_names]] for period in periods],
    )

    if not periods or not series_names:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Horas"
    chart.x_axis.title = "Datas"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    chart.x_axis.majorTickMark = "out"
    chart.height = 12
    chart.width = 26
    data_ref = Reference(sheet, min_col=2, max_col=1 + len(series_names), min_row=1, max_row=1 + len(periods))
    category_ref = Reference(sheet, min_col=1, min_row=2, max_row=1 + len(periods))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    sheet.add_chart(chart, "H2")


def _append_comparison_chart_sheet(sheet, title: str, value_header: str, rows: list[list]) -> None:
    _append_sheet(sheet, ["Projeto", value_header], rows)
    if not rows:
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = value_header
    chart.x_axis.title = "Projeto"
    chart.height = 12
    chart.width = 26
    data_ref = Reference(sheet, min_col=2, min_row=1, max_row=1 + len(rows))
    category_ref = Reference(sheet, min_col=1, min_row=2, max_row=1 + len(rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    sheet.add_chart(chart, "D2")


def _evolution_point_label(index: int, point: dict) -> str:
    imported_at = str(point.get("importedAt", ""))[:10]
    return f"{index + 1}a - {imported_at}"


def seconds_to_hhmmss(total_seconds: int) -> str:
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
